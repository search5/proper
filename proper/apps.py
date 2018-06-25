# -*- coding: utf-8 -*-
import sys
import os
from proper import util
import json
import re
import math
from openpyxl import Workbook

from flask import Flask, abort, g, render_template, request, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from webhelpers import paginate
from dateutil.parser import parse
from StringIO import StringIO
import datetime
from dateutil.relativedelta import relativedelta


app = Flask(__name__)

@app.route("/")
def proper_main():
    return render_template("proper_main.html")


@app.route("/", methods=["POST"])
def proper_login():
    return redirect("/home")


@app.route("/home")
def proper_home():
    return render_template("home.html")


@app.route("/pc")
def proper_pc():
    return render_template("resource/pc.html")


@app.route("/pc/<asset_number>")
def proper_pc_view(asset_number):
    return render_template("resource/pc_view.html")


@app.route("/monitor")
def proper_monitor():
    return render_template("resource/monitor.html")


@app.route("/monitor/<asset_number>")
def proper_monitor_view(asset_number):
    return render_template("resource/monitor_view.html")


@app.route("/printer")
def proper_printer():
    return render_template("resource/printer.html")


@app.route("/printer/<asset_number>")
def proper_printer_view(asset_number):
    return render_template("resource/printer_view.html")


@app.route("/equipment")
def proper_equipment():
    return render_template("resource/equipment.html")


@app.route("/equipment/<asset_number>")
def proper_equipment_view(asset_number):
    return render_template("resource/equipment_view.html")


@app.route("/sr")
def proper_sr():
    return render_template("service/sr.html")


@app.route("/sr/<service_number>")
def proper_sr_view(service_number):
    return render_template("service/sr_view.html")


@app.route("/<user>/use")
def user_use(user):
    return render_template("user/use.html")


@app.route("/<user>/modify")
def user_modify(user):
    return render_template("user/modify.html")


@app.route("/logout")
def user_logout():
    return ""

"""
app.config['UPLOAD_DIR'] = os.path.join(os.path.dirname(__file__), "uploads")
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql+psycopg2://proper:1234@localhost/proper'
db = SQLAlchemy(app)

from proper.models import User, Dept, ServiceRequest, Asset, AssetHistory

def oldYears(dob):
    today = datetime.date.today()

    if today.month < dob.month or \
      (today.month == dob.month and today.day < dob.day):
        return today.year - dob.year - 1
    else:
        return today.year - dob.year



@app.before_request
def before_request():
    remote_addr = request.remote_addr

    remote_first_addr = remote_addr[:remote_addr.rfind(".") + 1]

    # 접속 허용이 가능한 IP입니까?
    #if remote_first_addr not in ("127.0.0.", "121.131.33."):
    #    abort(403)


@app.route("/")
def main():
    return render_template("index.html")


@app.errorhandler(403)
def page_not_allow(e):
    return render_template('error/403.html'), 403


@app.route("/asset")
def asset_list():
    print type(request.stream)
    return render_template("asset/list.html")


@app.route("/pc", endpoint='pc')
def pc_list():
    page = int(request.args.get('page', 1))
    items_per_page = 100

    q = db.session.query(Asset).filter(Asset.asset_type == 'desktop').order_by(db.desc(Asset.seq))
    page_url = paginate.PageURL("/pc", {"page": page})
    records = paginate.Page(q, page, url=page_url, items_per_page=items_per_page)

    view_data = {
        "total": q.count(),
        "records": records
    }

    db.session.commit()

    return render_template("asset/pc_list.html", **view_data)

@app.route("/pc/create", methods=("POST",))
def pc_asset_create():
    pc_info = {}

    if request.files['file_xml']:
        new_xml_filename = os.path.join(app.config['UPLOAD_DIR'], util.generateTimeBaseFileName(request.files['file_xml'].filename))
        request.files['file_xml'].save(new_xml_filename)
    
        # XML 파일을 분석해서 정보 채워넣기
        pc_info = util.parsePCInfoXML(new_xml_filename)

    asset = Asset()
    asset.manufacture = pc_info['manufacture']
    asset.video_chip = pc_info['video_chip']
    asset.scan_time = parse(pc_info['scan_time'])
    asset.service_pack = pc_info['service_pack']
    asset.processor_version = pc_info['cpu_name']
    asset.hdd_name = pc_info['hdd_name']
    asset.memory_size = pc_info['memory_size']
    asset.hdd_capacity = pc_info['hdd_capacity']
    asset.hdd_serial = pc_info['hdd_serial']
    asset.product_serial = pc_info['product_serial']
    asset.mac_address = pc_info['mac_addr']
    asset.os = pc_info['os']
    asset.product_name = pc_info['product_name']
    asset.usb_support_type = pc_info['usb_support_ver']
    asset.asset_type = 'desktop'
    asset.xml_file_name = request.files['file_xml'].filename
    asset.xml_file_content = new_xml_filename

    db.session.add(asset)
    db.session.commit()

    asset.asset_code = util.makeAssetCode(asset.seq)
    db.session.commit()

    return redirect(url_for('pc'))

@app.route("/pc/create/empty", methods=("POST",))
def pc_asset_create_empty():
    asset = Asset()
    asset.asset_type = 'desktop'

    db.session.add(asset)
    db.session.commit()

    asset.asset_code = util.makeAssetCode(asset.seq)
    db.session.commit()

    return ""

@app.route("/pc/<asset_code>", endpoint="pc_view")
def pc_asset_view(asset_code):

    asset = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    asset_dict = asset.__dict__
    asset_dict['user'] = asset.user # Asset 레코드를 __dict__를 사용해 변환하는 경우 relationship은 상속받지 못한다.

    # 항목 변경내역 가져오기
    asset_history = db.session.query(AssetHistory).filter(AssetHistory.asset_code == asset_code)
    asset_dict['asset_history'] = asset_history

    #print dir(asset)

    return render_template("asset/pc_view.html", **asset_dict)

# 정보 업데이트
@app.route("/pc/<asset_code>/update", methods=("POST",))
def pc_asset_update(asset_code):
    user_input_dict = {
        'video_chip': request.values.get('video_chip', ''),
        'manufacture': request.values.get('manufacture', ''),
        'netbios_name': request.values.get('netbios_name', ''),
        'hdd_name': request.values.get('hdd_name', ''),
        'memory_size': request.values.get('memory_size', ''),
        'disuse_date': request.values.get('disuse_date', ''),
        'acquire_money': request.values.get('acquire_money', ''),
        'windows_serial': request.values.get('windows_serial', ''),
        'asset_code': request.values.get('asset_code', ''),
        'note': request.values.get('note', ''),
        'mac_address': request.values.get('mac_address', ''),
        'product_name': request.values.get('product_name', ''),
        'make_date': request.values.get('make_date', ''),
        'windows_version': request.values.get('windows_version', ''),
        'service_pack': request.values.get('service_pack', ''),
        'config_no': request.values.get('config_no', ''),
        'hdd_capacity': request.values.get('hdd_capacity', ''),
        'processor_version': request.values.get('processor_version', ''),
        'hdd_serial': request.values.get('hdd_serial', ''),
        'product_serial': request.values.get('product_serial', ''),
        'acquire_date': request.values.get('acquire_date', ''),
        'usb_support_type': request.values.get('usb_support_type', ''),
        'os': request.values.get('os', ''),
        'disuse_note': request.values.get('disuse_note', ''),
        'send_date': request.values.get('send_date', ''),
        'collect_date': request.values.get('collect_date', ''),
        'user_seq': request.values.get('user_seq', 0)
    }

    # 사용자가 입력한 것과 데이터베이스에 저장되어 있는 값을 마이너스 연산해서 어떤 값을 변경했는지 기록을 님기게 한다.

    # Database Record
    record = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    changed_data = record - user_input_dict

    # 변경내역이 None이 아니고 하나 이상일 경우 수행하게 한다.
    if changed_data != None:
        for item in changed_data:
            # 히스토리 레코드 생성
            asset_history = AssetHistory(asset_code, **item)
            db.session.add(asset_history)

    # 할당 정보 입력을 위해 행동을 취한다.
    user_seq = request.values.get('user_seq', 0)
    user_name = request.values.get('user_name', '')

    # 사용자 정보 할당
    user = db.session.query(User).filter(User.seq == user_seq).first()
    record.user = user

    db.session.commit()

    return redirect(url_for("pc_view", asset_code=asset_code))

# 정보 업데이트
@app.route("/pc/<asset_code>/updateXml", methods=("POST",))
def pc_asset_update_xml(asset_code):
    print asset_code
    if request.files['file_xml']:
        new_xml_filename = os.path.join(app.config['UPLOAD_DIR'], util.generateTimeBaseFileName(request.files['file_xml'].filename))
        request.files['file_xml'].save(new_xml_filename)

        # XML 파일을 분석해서 정보 채워넣기
        pc_info = util.parsePCInfoXML(new_xml_filename)

        asset = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
        asset.manufacture = pc_info['manufacture']
        asset.video_chip = pc_info['video_chip']
        asset.scan_time = parse(pc_info['scan_time'])
        asset.service_pack = pc_info['service_pack']
        asset.processor_version = pc_info['cpu_name']
        asset.hdd_name = pc_info['hdd_name']
        asset.memory_size = pc_info['memory_size']
        asset.hdd_capacity = pc_info['hdd_capacity']
        asset.hdd_serial = pc_info['hdd_serial']
        asset.product_serial = pc_info['product_serial']
        asset.mac_address = pc_info['mac_addr']
        asset.os = pc_info['os']
        asset.product_name = pc_info['product_name']
        asset.usb_support_type = pc_info['usb_support_ver']
        asset.xml_file_name = request.files['file_xml'].filename

        db.session.commit()

    return redirect(url_for("pc_view", asset_code=asset_code))

@app.route("/notebook", endpoint='notebook')
def notebook_list():
    page = int(request.args.get('page', 1))
    items_per_page = 100

    q = db.session.query(Asset).filter(Asset.asset_type == 'notebook').order_by(db.desc(Asset.seq))
    page_url = paginate.PageURL("/notebook", {"page": page})
    records = paginate.Page(q, page, url=page_url, items_per_page=items_per_page)

    view_data = {
        "total": q.count(),
        "records": records
    }

    db.session.commit()

    return render_template("asset/notebook_list.html", **view_data)

@app.route("/notebook/create", methods=("POST",))
def notebook_asset_create():
    pc_info = {}

    if request.files['file_xml']:
        new_xml_filename = os.path.join(app.config['UPLOAD_DIR'], util.generateTimeBaseFileName(request.files['file_xml'].filename))
        request.files['file_xml'].save(new_xml_filename)

        # XML 파일을 분석해서 정보 채워넣기
        pc_info = util.parsePCInfoXML(new_xml_filename)

    asset = Asset()
    asset.manufacture = pc_info['manufacture']
    asset.video_chip = pc_info['video_chip']
    asset.scan_time = parse(pc_info['scan_time'])
    asset.service_pack = pc_info['service_pack']
    asset.processor_version = pc_info['cpu_name']
    asset.hdd_name = pc_info['hdd_name']
    asset.memory_size = pc_info['memory_size']
    asset.hdd_capacity = pc_info['hdd_capacity']
    asset.hdd_serial = pc_info['hdd_serial']
    asset.product_serial = pc_info['product_serial']
    asset.mac_address = pc_info['mac_addr']
    asset.os = pc_info['os']
    asset.product_name = pc_info['product_name']
    asset.usb_support_type = pc_info['usb_support_ver']
    asset.asset_type = 'notebook'
    asset.xml_file_name = request.files['file_xml'].filename
    asset.xml_file_content = new_xml_filename

    db.session.add(asset)
    db.session.commit()

    asset.asset_code = util.makeAssetCode(asset.seq)
    db.session.commit()

    return redirect(url_for('notebook'))

@app.route("/notebook/<asset_code>", endpoint="notebook_view")
def notebook_asset_view(asset_code):

    asset = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    asset_dict = asset.__dict__
    asset_dict['user'] = asset.user # Asset 레코드를 __dict__를 사용해 변환하는 경우 relationship은 상속받지 못한다.

    # 항목 변경내역 가져오기
    asset_history = db.session.query(AssetHistory).filter(AssetHistory.asset_code == asset_code)
    asset_dict['asset_history'] = asset_history

    return render_template("asset/notebook_view.html", **asset_dict)

# 정보 업데이트
@app.route("/notebook/<asset_code>/update", methods=("POST",))
def notebook_asset_update(asset_code):
    user_input_dict = {
        'video_chip': request.values.get('video_chip', ''),
        'manufacture': request.values.get('manufacture', ''),
        'netbios_name': request.values.get('netbios_name', ''),
        'hdd_name': request.values.get('hdd_name', ''),
        'memory_size': request.values.get('memory_size', ''),
        'disuse_date': request.values.get('disuse_date', ''),
        'acquire_money': request.values.get('acquire_money', ''),
        'windows_serial': request.values.get('windows_serial', ''),
        'asset_code': request.values.get('asset_code', ''),
        'note': request.values.get('note', ''),
        'mac_address': request.values.get('mac_address', ''),
        'product_name': request.values.get('product_name', ''),
        'make_date': request.values.get('make_date', ''),
        'windows_version': request.values.get('windows_version', ''),
        'service_pack': request.values.get('service_pack', ''),
        'config_no': request.values.get('config_no', ''),
        'hdd_capacity': request.values.get('hdd_capacity', ''),
        'processor_version': request.values.get('processor_version', ''),
        'hdd_serial': request.values.get('hdd_serial', ''),
        'product_serial': request.values.get('product_serial', ''),
        'acquire_date': request.values.get('acquire_date', ''),
        'usb_support_type': request.values.get('usb_support_type', ''),
        'os': request.values.get('os', ''),
        'disuse_note': request.values.get('disuse_note', ''),
        'send_date': request.values.get('send_date', ''),
        'collect_date': request.values.get('collect_date', ''),
        'user_seq': request.values.get('user_seq', 0)
    }

    # 사용자가 입력한 것과 데이터베이스에 저장되어 있는 값을 마이너스 연산해서 어떤 값을 변경했는지 기록을 님기게 한다.

    # Database Record
    record = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    changed_data = record - user_input_dict

    # 변경내역이 None이 아니고 하나 이상일 경우 수행하게 한다.
    if changed_data != None:
        for item in changed_data:
            # 히스토리 레코드 생성
            asset_history = AssetHistory(asset_code, **item)
            db.session.add(asset_history)

    # 할당 정보 입력을 위해 행동을 취한다.
    user_seq = request.values.get('user_seq', 0)
    user_name = request.values.get('user_name', '')

    # 사용자 정보 할당
    user = db.session.query(User).filter(User.seq == user_seq).first()
    record.user = user

    db.session.commit()

    return redirect(url_for("notebook_view", asset_code=asset_code))

@app.route("/monitor", endpoint='monitor')
def monitor_list():
    page = int(request.args.get('page', 1))
    items_per_page = 100

    q = db.session.query(Asset).filter(Asset.asset_type == 'monitor').order_by(db.desc(Asset.seq))
    page_url = paginate.PageURL("/monitor", {"page": page})
    records = paginate.Page(q, page, url=page_url, items_per_page=items_per_page)

    view_data = {
        "total": q.count(),
        "records": records
    }

    db.session.commit()

    return render_template("asset/monitor_list.html", **view_data)

@app.route("/monitor/create", methods=("POST",))
def monitor_asset_create():
    monitors_info = {}

    if request.files['file_xml']:
        new_xml_filename = os.path.join(app.config['UPLOAD_DIR'], util.generateTimeBaseFileName(request.files['file_xml'].filename))
        request.files['file_xml'].save(new_xml_filename)

        # XML 파일을 분석해서 정보 채워넣기
        monitors_info = util.parseMonitorInfoXML(new_xml_filename)

    for monitor in monitors_info:
        date_parse = re.search('Week: ([0-9]{,2}), Year: ([0-9]{,4})', monitor['make_date'])
        asset = Asset()
        asset.product_serial = monitor['product_serial']
        asset.product_name = monitor['product_name']
        asset.asset_type = 'monitor'
        asset.make_date ="{1}-{0}".format(date_parse.group(1), date_parse.group(2))
        asset.xml_file_name = request.files['file_xml'].filename
        asset.xml_file_content = new_xml_filename

        db.session.add(asset)
        db.session.commit()

        asset.asset_code = util.makeAssetCode(asset.seq)
        db.session.commit()

    return redirect(url_for('monitor'))

@app.route("/monitor/<asset_code>", endpoint="monitor_view")
def monitor_asset_view(asset_code):

    asset = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    asset_dict = asset.__dict__
    asset_dict['user'] = asset.user # Asset 레코드를 __dict__를 사용해 변환하는 경우 relationship은 상속받지 못한다.

    # 항목 변경내역 가져오기
    asset_history = db.session.query(AssetHistory).filter(AssetHistory.asset_code == asset_code)
    asset_dict['asset_history'] = asset_history

    return render_template("asset/monitor_view.html", **asset_dict)

# 정보 업데이트
@app.route("/monitor/<asset_code>/update", methods=("POST",))
def monitor_asset_update(asset_code):
    user_input_dict = {
        'manufacture': request.values.get('manufacture', ''),
        'asset_code': request.values.get('asset_code', ''),
        'make_date': request.values.get('make_date', ''),
        'acquire_money': request.values.get('acquire_money', ''),
        'product_name': request.values.get('product_name', ''),
        'product_serial': request.values.get('product_serial', ''),
        'acquire_date': request.values.get('acquire_date', ''),
        'disuse_date': request.values.get('disuse_date', ''),
        'disuse_note': request.values.get('disuse_note', ''),
        'send_date': request.values.get('send_date', ''),
        'collect_date': request.values.get('collect_date', ''),
        'user_seq': request.values.get('user_seq', 0)
    }

    # 사용자가 입력한 것과 데이터베이스에 저장되어 있는 값을 마이너스 연산해서 어떤 값을 변경했는지 기록을 님기게 한다.

    # Database Record
    record = db.session.query(Asset).filter(Asset.asset_code == asset_code).first()
    changed_data = record - user_input_dict

    # 변경내역이 None이 아니고 하나 이상일 경우 수행하게 한다.
    if changed_data != None:
        for item in changed_data:
            # 히스토리 레코드 생성
            asset_history = AssetHistory(asset_code, **item)
            db.session.add(asset_history)

    # 할당 정보 입력을 위해 행동을 취한다.
    user_seq = request.values.get('user_seq', 0)

    # 사용자 정보 할당
    user = db.session.query(User).filter(User.seq == user_seq).first()
    record.user = user

    db.session.commit()

    return redirect(url_for("monitor_view", asset_code=asset_code))

@app.route("/user", endpoint="user")
def user_list():
    page = int(request.args.get('page', 1))
    items_per_page = 100

    q = db.session.query(User)
    total_cnt = q.count()
    q = q.order_by(db.desc(User.seq))
    page_url = paginate.PageURL("/user", {"page": page})
    records = paginate.Page(q, page, url=page_url, items_per_page=items_per_page)

    # 팀 목록 제공하기
    team_list = db.session.query(Dept)

    view_data = {
        "total": total_cnt,
        "records": records,
        "team_list": team_list
    }

    db.session.commit()

    return render_template("mngt/user.html", **view_data)

@app.route("/user/<user_seq>", endpoint="user_view")
def user_view(user_seq):
    q = db.session.query(User).filter(User.seq == user_seq).first()

    # 팀 목록 제공하기
    team_list = db.session.query(Dept)

    # 사용했던 자산목록 가져오기
    assets_history = db.session.query(AssetHistory.asset_code).filter(AssetHistory.field_name == 'user_seq')
    assets_history = assets_history.filter(AssetHistory.old_value == user_seq)

    assets_history = db.session.query(Asset).filter(Asset.asset_code.in_(assets_history))
    #print assets_history.all()

    view_data = q.__dict__
    view_data['assets'] = q.assets
    view_data['dept'] = q.dept
    view_data['team_list'] = team_list
    view_data['assets_history'] = assets_history

    return render_template("mngt/user_view.html", **view_data)

@app.route("/user/<user_seq>/update", methods=['POST'])
def user_update(user_seq):
    # dept, user_name, user_position이 넘어온다.
    dept = request.values.get('dept', 0)
    user_name = request.values.get('user_name', '')
    user_position = request.values.get('user_position', '')
    user_ip = request.values.get('user_ip', '')
    user_seq = request.values.get('user_seq', 0)

    dept_record = db.session.query(Dept).filter(Dept.seq == dept).first()

    update_user = db.session.query(User).filter(User.seq == user_seq).first()
    update_user.dept = dept_record
    update_user.user_name = user_name
    update_user.user_position = user_position
    update_user.user_ip = user_ip

    db.session.commit()

    return redirect(url_for("user", user_seq=user_seq))

@app.route("/user/create", methods=("POST",))
def user_create():
    # dept, user_name, user_position이 넘어온다.
    dept = request.values.get('dept', 0)
    user_name = request.values.get('user_name', '')
    user_position = request.values.get('user_position', '')
    user_ip = request.values.get('user_ip', '')

    dept_record = db.session.query(Dept).filter(Dept.seq == dept).first()

    new_user = User(dept_record, user_name, user_position, user_ip)

    db.session.add(new_user)
    db.session.commit()

    return redirect(url_for("user"))

@app.route("/user/all")
def user_all():
    # 사용자 목록을 모두 리스트로 만들어 사용자에게 보낸다.
    q = db.session.query(User)

    user_list = []
    for user in q:
        user_detail = {
            'user_seq': user.seq,
            'user_name': '%s %s' % (user.dept.dept_team_nm, user.user_name)
        }
        user_list.append(user_detail)

    return json.dumps(user_list)


@app.route("/team", endpoint="team")
def team_list():
    records = db.session.query(Dept)

    view_data = {
        "total": records.count(),
        "records": records
    }

    db.session.commit()

    return render_template("mngt/team.html", **view_data)

@app.route("/team/create", methods=("POST",))
def team_create():
    team_name = request.values.get('team_name', '')
    if team_name != "":
        # 팀 이름이 들어왔을때만 팀을 생성하게 한다.
        team = Dept(team_name)

        db.session.add(team)
        db.session.commit()

    return redirect(url_for("team"))

@app.route("/team/<seq>/update", methods=("POST",))
def team_update(seq):
    team_seq = request.values.get('pk', 0)
    team_name = request.values.get('value', '')

    if team_name != "":
        # 팀 이름이 들어왔을때만 팀을 업데이트하게 한다.
        team = db.session.query(Dept).filter(Dept.seq == team_seq).first()
        team.dept_team_nm = team_name

        db.session.commit()

    return redirect(url_for("team"))

@app.teardown_request
def teardown_request(exception):
    pass

@app.route("/sr")
def sr():
    # 접속된 IP를 확인하여 사용자 이름 및 현재 사용중인 어셋 정보 등을 채워준다.
    connect_user_ip = request.remote_addr
    connect_user = db.session.query(User).filter(User.user_ip == connect_user_ip).first()

    view_data = {
        'user_info': connect_user
    }

    return render_template("sr/list.html", **view_data)

########################################## Excel
@app.route("/pc/excel")
def pc_excel():
    pass

@app.route("/monitor/execel")
def monitor_excel():
    pass

@app.route("/notebook/excel")
def notebook_excel():
    pass

@app.route("/userAsset/excel")
def user_asset_excel():
    q = db.session.query(User).order_by(db.desc(User.seq))

    output = StringIO()

    # 컬럼 : 자산구분/부서/이용자/제품명/제조년월/프로세서(PC)/메모리(PC)/비고/사용년수/제조년도/OS
    wb = Workbook()
    ws = wb.get_active_sheet()

    # sheet setting
    ws.title = "Sheet1"

    # db to excel
    column_header = [
        {
            "title": u"자산구분",
            "db_column": "asset_type",
            "index": 0
        },
        {
            "title": u"부서",
            "db_column": "user",
            "sub_path": "dept.dept_team_nm",
            "index": 1
        },
        {
            "title": u"이용자",
            "db_column": "user",
            "sub_path": "user_name",
            "index": 2
        },
        {
            "title": u"제품명",
            "db_column": "product_name",
            "index": 3
        },
        {
            "title": u"제조년월",
            "db_column": "make_date",
            "index": 4
        },
        {
            "title": u"프로세서(PC)",
            "db_column": "processor_version",
            "index": 5
        },
        {
            "title": u"메모리(PC)",
            "db_column": "memory_size",
            "index": 6
        },
        {
            "title": u"비고",
            "db_column": "note",
            "index": 7
        },
        {
            "title": u"사용년수",
            "db_column": "make_date",
            "index": 8
        },
        {
            "title": u"제조년도",
            "db_column": "make_date",
            "index": 9
        },
        {
            "title": u"OS",
            "db_column": "os",
            "index": 10
        }
    ]

    # excel column title setting(
    for column in column_header:
        cell = ws.cell(row=0, column=column["index"])
        cell.value = column["title"]

    def sheet_row_writing(ws, row, row_data):
        for column in column_header:
            cell = ws.cell(row=row, column=column["index"])

            db_column = column["db_column"]
            if db_column:
                value = getattr(row_data, db_column)

                sub_path = column.get('sub_path', '')
                if sub_path:
                    # sub_path가 2단계인가?
                    if "." in sub_path:
                        parent_value = getattr(value, sub_path.split(".")[0])
                        value = getattr(parent_value, sub_path.split(".")[1])
                        #value = sub_path.split(".")[1]
                    else:
                        value = getattr(value, sub_path)

                asset_type = {
                    'desktop': '데스크탑',
                    'notebook': '노트북',
                    'monitor': '모니터'
                }


                if db_column == "asset_type":
                    value = asset_type[value]
                elif db_column == "make_date":
                    if value:
                        tmp_date = re.split("[\.\-]", value)
                        # tmp_date의 월이 12이상인 경우 12로 강제 조정한다.
                        if tmp_date[1] > 12:
                            tmp_date[1] = 12
                        value = datetime.date(int(tmp_date[0]), int(tmp_date[1]), 1)

                if column["title"] == u"제조년월":
                    if type(value) == datetime.date:
                        value = value.month
                elif column["title"] == u"제조년도":
                    if type(value) == datetime.date:
                        value = value.year
                elif column["title"] == u"사용년수":
                    #if value and "." in value:
                    #    value = value.split(".")[0]
                    today = datetime.datetime.today().date()
                    if value:
                        #today_delta = today - value
                        value = oldYears(value) + 1
                    else:
                        value = 0

                cell.value = value or ''

    row = 1
    for user_entry in q:
        user_asset = user_entry.assets
        for user_asset_entry in user_asset:
            sheet_row_writing(ws, row, user_asset_entry)
            row += 1

    wb.save(output)
    output.seek(0)

    db.session.commit()

    #return render_template("mngt/user.html", **view_data)
    return send_file(output, mimetype="application/zip", as_attachment=True,
                     attachment_filename="AssetList.xlsx")
"""
